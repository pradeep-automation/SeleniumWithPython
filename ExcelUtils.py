import openpyxl

from openpyxl.styles import PatternFill


def get_row_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def get_column_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def read_data(file, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.cell(rownum, colnum).value)


def write_data(file, sheetName, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, colnum).value = data
    workbook.save(file)









