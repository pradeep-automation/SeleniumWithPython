import openpyxl
from selenium import webdriver

file = "C:/Development/robot-scripts/SeleniumWithPython/Data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row
cols = sheet.max_column

print(rows, cols)

# Reading all the rows and columns from the excel sheet

for r in range(2, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value, end= "---------")
    print()

